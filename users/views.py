from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import logout as auth_logout
from django.contrib.auth import authenticate, login
from django.contrib import messages
from decimal import Decimal
from django.utils.timezone import now
from users.utils import credit_profit
from django.http import HttpResponse, JsonResponse   # merged HttpResponse + JsonResponse
from django.db.models import Sum, Count              # merged Sum + Count
from datetime import date, timedelta
from django.utils import timezone
from decimal import Decimal,InvalidOperation
from django.core.paginator import Paginator
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas                  # corrected placement with other reportlab imports
from crypto.models import Wallet, InvestmentPlan, MarketInstrument
from django.db import transaction
import openpyxl
from openpyxl.utils import get_column_letter
import feedparser
import logging
from django.urls import reverse
from django.contrib.auth.models import User
from django.contrib.admin.views.decorators import staff_member_required
import json
import requests
import base64
import io
import pyotp
import qrcode
import csv
import re
from .models import (
    UserBalance,
    ActiveInvestment,
    Referral,
    UserKYC,
    Notification,
    UserWallet,
    SupportArticle,
    UserVerification,
    RecoveryCode,
    APIKey,
    UserProfile,
    Transaction,
    Deposit,
    Withdrawal,
    ProfitRecord,
    CompanyWallet,
    ReferralCommission
)
from.models import CompanyWallet
from pycoingecko import CoinGeckoAPI
cg = CoinGeckoAPI()

from .forms import (
    KYCForm,
    UserWalletForm,
    ProfileForm,
    NotificationsForm,
    AccountForm,
    VerificationForm,
    APIForm,
    SettingsForm,
    WithdrawalForm
)

logger = logging.getLogger(__name__)
MIN_WITHDRAWAL = Decimal("200.00") 


@login_required
def dashboard(request):
    user = request.user

    instruments = MarketInstrument.objects.all()

    profile = UserProfile.objects.get(user=user)

    investments = ActiveInvestment.objects.filter(user=user)

    referral, _ = Referral.objects.get_or_create(user=user)

    notifications = Notification.objects.filter(
        user=user
    ).order_by("-timestamp")[:10]


    # ==================================================
    # AUTO-COMPLETE MATURED INVESTMENTS
    # ==================================================

    for inv in investments.filter(status="active"):

        if timezone.now() >= inv.end_date:

            inv.status = "completed"
            inv.save()


    # ==================================================
    # REFRESH INVESTMENTS AFTER STATUS CHANGES
    # ==================================================

    investments = ActiveInvestment.objects.filter(user=user)


    # ==================================================
    # GROWTH CHART DATA
    # ==================================================

    growth_data = []

    growth_labels = []

    total = 0


    for inv in investments.order_by("start_date"):

        total += float(inv.amount)

        growth_data.append(round(total, 2))

        growth_labels.append(
            inv.start_date.strftime("%b %d")
        )


    # ==================================================
    # ROI CALCULATIONS
    # ==================================================

    daily_roi = 0

    weekly_roi = 0

    monthly_roi = 0


    for inv in investments:

        days = (
            inv.end_date - inv.start_date
        ).days


        if days > 0:

            daily_roi += (
                float(inv.amount) / days
            )

            weekly_roi += (
                float(inv.amount) / days
            ) * 7

            monthly_roi += (
                float(inv.amount) / days
            ) * 30


    roi_data = [

        round(daily_roi, 2),

        round(weekly_roi, 2),

        round(monthly_roi, 2),

    ]


    # ==================================================
    # EXTERNAL MARKET DATA
    # ==================================================

    try:

        crypto_url = (
            "https://api.coingecko.com/api/v3/"
            "simple/price"
            "?ids=bitcoin,ethereum"
            "&vs_currencies=usd"
        )

        crypto_response = requests.get(
            crypto_url,
            timeout=10
        )

        crypto_data = crypto_response.json()


        btc_price = crypto_data[
            "bitcoin"
        ]["usd"]


        eth_price = crypto_data[
            "ethereum"
        ]["usd"]


        fmp_url = (
            "https://financialmodelingprep.com/api/v3/"
            "quote/%5EGSPC"
            "?apikey=ZR8hDHF0vtAETUxVCfT41Du5Wtc9fzEO"
        )


        fmp_response = requests.get(
            fmp_url,
            timeout=10
        )

        sp500_data = fmp_response.json()


        sp500_index = sp500_data[0]["price"]


    except Exception:

        btc_price = "N/A"

        eth_price = "N/A"

        sp500_index = "N/A"


    # ==================================================
    # REFERRALS
    # ==================================================

    referred_users = User.objects.filter(
        referral__referrer=user
    )

    referral_count = referred_users.count()


    # ==================================================
    # REFERRAL BONUS
    # ==================================================

    referral_bonus = (
        ReferralCommission.objects
        .filter(referrer=user)
        .aggregate(
            total=Sum("commission_amount")
        )["total"]
        or 0
    )


    # ==================================================
    # COMPLETED INVESTMENTS
    # ==================================================

    completed_investments = (
        investments.filter(
            status="completed"
        )
    )


    total_profit = sum(

        [
            inv.get_current_value() - inv.amount

            for inv in completed_investments
        ]

    )


    # ==================================================
    # TOTAL BALANCE
    # ==================================================

    total_balance = (

        profile.usd_balance

        + profile.investment_balance

        + profile.profit_balance

        + profile.bonus_balance

        + sum(
            [
                inv.get_current_value()

                for inv in completed_investments
            ]
        )

    )


    # ==================================================
    # DASHBOARD CONTEXT
    # ==================================================

    context = {

        "instruments": instruments,

        "profile": profile,

        "investments": investments,

        "referral": referral,

        "referral_count": referral_count,

        "referral_bonus": referral_bonus,

        # IMPORTANT:
        # Pass Python lists directly.
        # json_script handles the JSON conversion.

        "growth_data": growth_data,

        "growth_labels": growth_labels,

        "roi_data": roi_data,

        "notifications": notifications,

        "btc_price": btc_price,

        "eth_price": eth_price,

        "sp500_index": sp500_index,

        "total_profit": round(
            total_profit,
            2
        ),

        "total_balance": round(
            total_balance,
            2
        ),

        "investment_balance":
            profile.investment_balance,

    }


    return render(
        request,
        "users/dashboard.html",
        context
    )

@login_required
def portfolio(request):
    investments = ActiveInvestment.objects.filter(user=request.user)
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')[:5]
    total_invested = investments.aggregate(Sum('amount'))['amount__sum'] or Decimal("0")
    current_value = sum(inv.get_current_value() for inv in investments)
    roi = ((current_value - total_invested) / total_invested * Decimal("100")) if total_invested else Decimal("0")

    best_asset = None
    if investments.exists():
        best_asset = max(investments, key=lambda inv: inv.get_current_multiplier()).plan_name
    upcoming_investments = investments.filter(end_date__isnull=False).order_by('end_date')[:5]
    prices = {}
    try:
        response = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=bitcoin,ethereum,tether&vs_currencies=usd")
        if response.status_code == 200:
            data = response.json()
            prices = {
                "BTC": data["bitcoin"]["usd"],
                "ETH": data["ethereum"]["usd"],
                "USDT": data["tether"]["usd"],
            }
    except Exception:
        prices = {"BTC": 0, "ETH": 0, "USDT": 0}

    risk_level = "Moderate"
    diversification_score = 75
    volatility_index = 12.5

    return render(request, "users/portfolio.html", {
        "holdings": json.dumps({k: v for k, v in prices.items()}),
        "investments": investments,
        "transactions": transactions,
        "total_invested": round(total_invested, 2),
        "current_value": round(current_value, 2),
        "roi": round(roi, 2),
        "best_asset": best_asset or "N/A",
        "risk_level": risk_level,
        "diversification_score": diversification_score,
        "volatility_index": volatility_index,
        "btc_price": prices.get("BTC"),
        "eth_price": prices.get("ETH"),
        "usdt_price": prices.get("USDT"),
        "upcoming_investments": upcoming_investments,
    })


@login_required
def investments(request):
    investments = ActiveInvestment.objects.filter(user=request.user)
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')[:10]
    total_invested = (
        investments.aggregate(Sum('amount'))['amount__sum']
        or Decimal("0")
    )
    current_value = sum(
        (inv.get_current_value() for inv in investments),
        Decimal("0")
    )

    if total_invested > Decimal("0"):
        roi = (
            (current_value - total_invested)
            / total_invested
        ) * Decimal("100")
    else:
        roi = Decimal("0")

    best_asset = (
        investments.first().plan_name
        if investments.exists()
        else "N/A"
    )

    upcoming_investments = (
        investments
        .filter(end_date__isnull=False)
        .order_by("end_date")[:5]
    )
    prices = {
        "BTC": 0,
        "ETH": 0,
        "USDT": 0,
    }

    try:
        response = requests.get(
            "https://api.coingecko.com/api/v3/simple/price"
            "?ids=bitcoin,ethereum,tether&vs_currencies=usd",
            timeout=10,
        )

        if response.status_code == 200:
            data = response.json()

            prices = {
                "BTC": data.get("bitcoin", {}).get("usd", 0),
                "ETH": data.get("ethereum", {}).get("usd", 0),
                "USDT": data.get("tether", {}).get("usd", 0),
            }

    except Exception:
        pass

    context = {
        "investments": investments,
        "transactions": transactions,
        "total_invested": round(total_invested, 2),
        "current_value": round(current_value, 2),
        "roi": round(roi, 2),
        "best_asset": best_asset,
        "upcoming_investments": upcoming_investments,
        "btc_price": prices["BTC"],
        "eth_price": prices["ETH"],
        "usdt_price": prices["USDT"],
        "holdings": json.dumps(prices),
    }

    return render(request, "users/investments.html", context)

@login_required
def investment_plans(request):
    plans = InvestmentPlan.objects.all()
    recommended_plan = None

    balance_record = UserBalance.objects.filter(user=request.user).first()
    balance = balance_record.balance if balance_record else Decimal("0.00")

    matching_plans = [
        plan for plan in plans
        if plan.min_amount <= balance <= plan.max_amount
    ]

    if matching_plans:
        recommended_plan = max(
            matching_plans,
            key=lambda plan: plan.roi_percent
        )
    elif balance > 0:
        below = [plan for plan in plans if plan.max_amount < balance]
        above = [plan for plan in plans if plan.min_amount > balance]

        if below:
            recommended_plan = max(
                below,
                key=lambda plan: plan.max_amount
            )
        elif above:
            recommended_plan = min(
                above,
                key=lambda plan: plan.min_amount
            )

    return render(
        request,
        "users/investment_plans.html",
        {
            "plans": plans,
            "recommended_plan": recommended_plan,
        },
    )

@login_required
def wallets(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)

    user_wallet, _ = UserWallet.objects.get_or_create(user=request.user)

    if request.method == "POST":
        user_wallet.btc_wallet = request.POST.get("btc_address")
        user_wallet.eth_wallet = request.POST.get("eth_address")
        user_wallet.usdt_erc20_wallet = request.POST.get("usdt_erc20_address")
        user_wallet.usdt_trc20_wallet = request.POST.get("usdt_trc20_address")
        user_wallet.save()

    recent_transactions = (
        Transaction.objects.filter(user=request.user)
        .order_by("-date")[:5]
    )

    notifications = (
        Notification.objects.filter(user=request.user)
        .order_by("-timestamp")[:5]
    )

    balances = {
        "USD": profile.usd_balance,
        "BTC": profile.btc_balance,
        "ETH": profile.eth_balance,
        "USDT ERC20": profile.usdt_erc20_balance,
        "USDT TRC20": profile.usdt_trc20_balance,
    }

    chart_data = (
        Transaction.objects.filter(user=request.user)
        .values("type")
        .annotate(count=Count("id"))
        .order_by("type")
    )

    wallet = Wallet.objects.filter(user=request.user).first()

    return render(
        request,
        "users/wallets.html",
        {
            "wallet": wallet,
            "user_wallet": user_wallet,
            "balances": balances,
            "recent_transactions": recent_transactions,
            "notifications": notifications,
            "chart_data": chart_data,

            "btc_address": user_wallet.btc_wallet,
            "eth_address": user_wallet.eth_wallet,
            "usdt_erc20_address": user_wallet.usdt_erc20_wallet,
            "usdt_trc20_address": user_wallet.usdt_trc20_wallet,
        },
    )

@login_required
def download_receipt(request, tx_id):
    tx = get_object_or_404(Transaction, id=tx_id, user=request.user)
    context = {
        "transaction": tx,
    }
    return render(request, "users/receipt.html", context)
    


@login_required
def withdraw(request):
    profile = UserProfile.objects.get(user=request.user)
    userwallet = UserWallet.objects.filter(user=request.user).first()

    wallets = {
        "BTC": userwallet.btc_wallet if userwallet else "",
        "ETH": userwallet.eth_wallet if userwallet else "",
        "USDT_ERC20": userwallet.usdt_erc20_wallet if userwallet else "",
        "USDT_TRC20": userwallet.usdt_trc20_wallet if userwallet else "",
    }

    form = WithdrawalForm()

    return render(
        request,
        "users/withdraw.html",   # ✅ always use withdraw.html
        {
            "form": form,
            "profile": profile,
            "userwallet": userwallet,
            "wallets": wallets,
        }
    )



@login_required
def transactions(request):
    qs = Transaction.objects.filter(user=request.user).order_by("-date")

    # Filters
    tx_type = request.GET.get("type")
    status = request.GET.get("status")
    search = request.GET.get("search")

    if tx_type:
        qs = qs.filter(type=tx_type)
    if status:
        qs = qs.filter(status=status)
    if search:
        qs = qs.filter(asset__icontains=search) | qs.filter(transaction_id__icontains=search)

    total_deposits = qs.filter(type="deposit").aggregate(Sum("amount"))["amount__sum"] or 0
    total_withdrawals = qs.filter(type="withdrawal").aggregate(Sum("amount"))["amount__sum"] or 0
    net_change = total_deposits - total_withdrawals
    total_transactions = qs.count()

    paginator = Paginator(qs, 10)  
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "transactions": page_obj,
        "total_deposits": total_deposits,
        "total_withdrawals": total_withdrawals,
        "net_change": net_change,
        "total_transactions": total_transactions,
    }
    return render(request, "users/transactions.html", context)

@login_required
def export_transactions_csv(request):
    qs = Transaction.objects.filter(user=request.user).order_by("-date")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="transactions.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Type", "Asset", "Amount", "Status"])
    for tx in qs:
        writer.writerow([tx.date, tx.type, tx.asset, tx.amount, tx.status])

    return response

@login_required
def export_transactions_pdf(request):
    qs = Transaction.objects.filter(user=request.user).order_by("-date")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="transactions.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica", 12)

    y = 750
    p.drawString(100, y, "Transactions Report")
    y -= 30

    for tx in qs:
        line = f"{tx.date} | {tx.type} | {tx.asset} | {tx.amount} | {tx.status}"
        p.drawString(100, y, line)
        y -= 20

    p.showPage()
    p.save()
    return response

@login_required
def export_transactions_excel(request):
    qs = Transaction.objects.filter(user=request.user).order_by("-date")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Type", "Asset", "Amount", "Status"]
    ws.append(headers)

    for tx in qs:
        ws.append([tx.date, tx.type, tx.asset, tx.amount, tx.status])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="transactions.xlsx"'
    wb.save(response)
    return response


@login_required
def referrals(request):
    user = request.user

    referral, _ = Referral.objects.get_or_create(user=user)

    referral_link = (
        f"{request.scheme}://{request.get_host()}"
        f"{reverse('crypto:signup')}?ref={user.username}"
    )

    profile, _ = UserProfile.objects.get_or_create(user=user)

    referral_earnings = getattr(profile, "referral_earnings", 0)

    referred_users = User.objects.filter(referral__referrer=user)

    referral_count = referred_users.count()

    return render(
        request,
        "users/referrals.html",
        {
            "referral_link": referral_link,
            "referral_count": referral_count,
            "referral_earnings": referral_earnings,
            "referred_users": referred_users,
        },
    )

@login_required
def markets(request):
    return render(request, "users/markets.html")

# Simple cache to store last good prices
last_assets = {}

@login_required
def market_data(request):
    # --- Prices ---
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "ids": "bitcoin,ethereum,tether"
    }
    try:
        response = requests.get(url, params=params).json()
        # If CoinGecko returns an error (rate limit), reuse cached data
        if isinstance(response, dict) and "status" in response:
            logger.warning(f"CoinGecko error: {response}")
            assets = last_assets or {
                "BTC": {"price": None, "change": None},
                "ETH": {"price": None, "change": None},
                "USDT_ERC20": {"price": None, "change": None},
                "USDT_TRC20": {"price": None, "change": None},
            }
        else:
            # Map response into assets dict
            assets = {}
            for coin in response:
                if coin["id"] == "bitcoin":
                    assets["BTC"] = {"price": coin["current_price"], "change": coin["price_change_percentage_24h"]}
                elif coin["id"] == "ethereum":
                    assets["ETH"] = {"price": coin["current_price"], "change": coin["price_change_percentage_24h"]}
                elif coin["id"] == "tether":
                    assets["USDT_ERC20"] = {"price": coin["current_price"], "change": coin["price_change_percentage_24h"]}
                    assets["USDT_TRC20"] = {"price": coin["current_price"], "change": coin["price_change_percentage_24h"]}
            last_assets.update(assets)  # cache good data
    except Exception as e:
        logger.error(f"CoinGecko price API failed: {e}")
        assets = last_assets or {
            "BTC": {"price": None, "change": None},
            "ETH": {"price": None, "change": None},
            "USDT_ERC20": {"price": None, "change": None},
            "USDT_TRC20": {"price": None, "change": None},
        }

    # --- Trending ---
    try:
        trending_url = "https://api.coingecko.com/api/v3/search/trending"
        trending_data = requests.get(trending_url).json()
        trending = [coin["item"]["symbol"] for coin in trending_data.get("coins", [])]
    except Exception as e:
        logger.error(f"Trending API failed: {e}")
        trending = []

    # --- Chart (7 days BTC) ---
    chart_url = "https://api.coingecko.com/api/v3/coins/bitcoin/market_chart"
    chart_params = {"vs_currency": "usd", "days": "7"}
    try:
        chart_data = requests.get(chart_url, params=chart_params).json()
        labels = [f"Day {i+1}" for i in range(len(chart_data.get("prices", [])))]
        prices = [p[1] for p in chart_data.get("prices", [])]
    except Exception as e:
        logger.error(f"Chart API failed: {e}")
        labels, prices = [], []
    chart = {"labels": labels, "prices": prices}

    # --- Global overview ---
    try:
        global_url = "https://api.coingecko.com/api/v3/global"
        global_data = requests.get(global_url).json().get("data", {})
        overview = {
            "market_cap": global_data.get("total_market_cap", {}).get("usd"),
            "volume": global_data.get("total_volume", {}).get("usd"),
            "btc_dominance": global_data.get("market_cap_percentage", {}).get("btc"),
            "trending": trending
        }
    except Exception as e:
        logger.error(f"Global API failed: {e}")
        overview = {"market_cap": None, "volume": None, "btc_dominance": None, "trending": []}

    # --- Movers ---
    try:
        movers_url = "https://api.coingecko.com/api/v3/coins/markets"
        movers_params = {
            "vs_currency": "usd",
            "order": "market_cap_desc",
            "per_page": 10,
            "page": 1,
            "price_change_percentage": "24h"
        }
        movers_data = requests.get(movers_url, params=movers_params).json()
        sorted_coins = sorted(movers_data, key=lambda x: x.get("price_change_percentage_24h", 0))
        losers = sorted_coins[:3]
        gainers = sorted_coins[-3:]
        movers = {
            "gainers": [{"name": c["name"], "symbol": c["symbol"], "price": c["current_price"], "change": c["price_change_percentage_24h"]} for c in gainers],
            "losers": [{"name": c["name"], "symbol": c["symbol"], "price": c["current_price"], "change": c["price_change_percentage_24h"]} for c in losers]
        }
    except Exception as e:
        logger.error(f"Movers API failed: {e}")
        movers = {"gainers": [], "losers": []}

    # --- News ---
    try:
        feed = feedparser.parse("https://www.coindesk.com/arc/outboundfeeds/rss/")
        news = [{"title": entry.title, "url": entry.link} for entry in feed.entries[:5]]
    except Exception as e:
        logger.error(f"RSS feed failed: {e}")
        news = []

    return JsonResponse({
        "assets": assets,
        "chart": chart,
        "overview": overview,
        "movers": movers,
        "watchlist": [],
        "news": news
    })

@login_required
def news(request):
    try:
        # Fetch BTC and ETH prices
        btc_data = requests.get(
            "https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=usd"
        ).json()
        eth_data = requests.get(
            "https://api.coingecko.com/api/v3/simple/price?ids=ethereum&vs_currencies=usd"
        ).json()
        global_data = requests.get(
            "https://api.coingecko.com/api/v3/global"
        ).json()

        btc_price = btc_data["bitcoin"]["usd"]
        eth_price = eth_data["ethereum"]["usd"]
        market_cap = global_data["data"]["total_market_cap"]["usd"]

    except Exception as e:
        logger.error(f"Market data fetch failed: {e}")
        btc_price = eth_price = market_cap = "N/A"

    return render(request, "users/news.html", {
        "btc_price": btc_price,
        "eth_price": eth_price,
        "market_cap": market_cap,
    })


@login_required
def news_data(request):
    try:
        feed = feedparser.parse("https://www.coindesk.com/arc/outboundfeeds/rss/")
        articles = [
            {"title": entry.title, "url": entry.link, "published": entry.published}
            for entry in feed.entries[:10]
        ]
    except Exception as e:
        logger.error(f"RSS feed failed: {e}")
        articles = []
    return JsonResponse({"articles": articles})

@login_required
def support(request):
    articles = SupportArticle.objects.all().order_by("-created_at")
    return render(request, "users/support.html", {
        "articles": articles
    })


@login_required
def profile(request):
    user = request.user
    wallet, _ = UserWallet.objects.get_or_create(user=user)
    wallet_form = UserWalletForm(request.POST or None, instance=wallet)

    for field_name in ['btc_wallet', 'eth_wallet', 'usdt_erc20_wallet', 'usdt_trc20_wallet']:
        field = wallet_form.fields[field_name]
        if getattr(wallet, field_name):
            field.widget.attrs['readonly'] = 'readonly'
        field.widget.attrs['id'] = f'id_{field_name}'

    if request.method == "POST" and wallet_form.is_valid():
        wallet_form.save()
        return redirect("/users/profile#wallets")

    active_investments = ActiveInvestment.objects.filter(user=user, status="active")
    completed_investments = ActiveInvestment.objects.filter(user=user, status="completed")

    active_investments_count = active_investments.count()
    total_invested = ActiveInvestment.objects.filter(user=user).aggregate(total=Sum("amount"))["total"] or 0
    total_roi = sum(inv.get_current_value() - inv.amount for inv in completed_investments)

  
    recent_investments = ActiveInvestment.objects.filter(user=user).order_by("-start_date")[:5]

    referral_earnings = getattr(user.userprofile, "referral_earnings", 0)

  
    chart_labels = [inv.start_date.strftime("%Y-%m-%d") for inv in recent_investments]
    chart_data = [float(inv.get_current_value()) for inv in recent_investments]

    return render(request, "users/profile.html", {
        "wallet_form": wallet_form,
        "active_investments_count": active_investments_count,
        "total_invested": total_invested,
        "total_roi": total_roi,
        "recent_investments": recent_investments,
        "referral_earnings": referral_earnings,
        "chart_labels": chart_labels,
        "chart_data": chart_data,
    })

@login_required
def profile_settings(request):
    user = request.user

    profile, _ = UserProfile.objects.get_or_create(user=user)
    kyc, _ = UserKYC.objects.get_or_create(user=user)
    verification, _ = UserVerification.objects.get_or_create(user=user)


    profile_form = ProfileForm(request.POST or None, request.FILES or None, instance=profile)
    notifications_form = NotificationsForm(request.POST or None, instance=profile)
    account_form = AccountForm(request.POST or None, instance=profile)
    kyc_form = KYCForm(request.POST or None, request.FILES or None, instance=kyc)
    verification_form = VerificationForm(request.POST or None, instance=verification)
    api_form = APIForm(request.POST or None)

    if request.method == "POST":
        if profile_form.is_valid(): profile_form.save()
        if notifications_form.is_valid(): notifications_form.save()
        if account_form.is_valid(): account_form.save()
        if kyc_form.is_valid(): kyc_form.save()
        if verification_form.is_valid(): verification_form.save()
        if api_form.is_valid():
            api_instance = api_form.save(commit=False)
            api_instance.user = user
            api_instance.save()
        return redirect("users:profile_settings")

    return render(request, "users/profile_settings.html", {
        "profile_form": profile_form,
        "notifications_form": notifications_form,
        "account_form": account_form,
        "kyc_form": kyc_form,
        "verification_form": verification_form,
        "api_form": api_form,
    })


@login_required
def quick_settings(request):
    user = request.user
    profile_form = ProfileForm(request.POST or None, request.FILES or None, instance=user.userprofile)
    notifications_form = NotificationsForm(request.POST or None, instance=user.userprofile)
    account_form = AccountForm(request.POST or None, instance=user.userprofile)
    kyc_form = KYCForm(request.POST or None, request.FILES or None, instance=user.userkyc)
    wallet_form = UserWalletForm(request.POST or None, instance=user.userwallet)
    verification_form = VerificationForm(request.POST or None, instance=user.userverification)
    api_form = APIForm(request.POST or None)

    if request.method == "POST":
        if profile_form.is_valid(): profile_form.save()
        if notifications_form.is_valid(): notifications_form.save()
        if account_form.is_valid(): account_form.save()
        if kyc_form.is_valid(): kyc_form.save()
        if wallet_form.is_valid(): wallet_form.save()
        if verification_form.is_valid(): verification_form.save()
        if api_form.is_valid():
            api_instance = api_form.save(commit=False)
            api_instance.user = user
            api_instance.save()
        return redirect("users:quick_settings")

    return render(request, "users/quick_settings.html", {
        "profile_form": profile_form,
        "notifications_form": notifications_form,
        "account_form": account_form,
        "kyc_form": kyc_form,
        "wallet_form": wallet_form,
        "verification_form": verification_form,
        "api_form": api_form,
    })



@login_required
def security(request):
    return render(request, "users/security.html")


@login_required
def settings(request):
    if request.method == "POST":
        form = SettingsForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Your settings have been updated successfully!")
            return redirect("users:settings")
    else:
        form = SettingsForm(instance=request.user)

    return render(request, "users/settings.html", {"form": form})


def login_alerts(request):
    return render(request, "users/login_alerts.html")


def view_activity(request):
    return render(request, "users/view_activity.html")


def revoke_sessions(request):
    return render(request, "users/revoke_sessions.html")


def change_password(request):
    return render(request, "users/change_password.html")


def manage_api_keys(request):
    return render(request, "users/manage_api_keys.html")


def security_alerts(request):
    return render(request, "users/security_alerts.html")


@login_required
def kyc_upload(request):
    kyc, created = UserKYC.objects.get_or_create(user=request.user)
    if request.method == "POST":
        form = KYCForm(request.POST, request.FILES, instance=kyc)
        if form.is_valid():
            form.save()
            kyc.status = "pending"
            kyc.save()
            messages.success(request, "Your KYC has been submitted for review.")
            return redirect("users:dashboard")
        else:
            messages.error(request, "There was a problem with your submission. Please check the form and try again.")
    else:
        form = KYCForm(instance=kyc)

    return render(request, "users/kyc_upload.html", {
        "form": form,
        "kyc": kyc
    })


@login_required
def logout_view(request):
    auth_logout(request)
    return redirect("crypto:login")


MIN_DEPOSIT = Decimal("200.00")

@login_required
def deposit(request):
    user = request.user

    profile = UserProfile.objects.get(user=user)

    wallet = UserWallet.objects.filter(user=user).first()

    wallet_exists = wallet is not None

    # Check whether the user has entered at least one
    # actual wallet address.
    wallet_has_address = False

    if wallet:
        wallet_has_address = any([
            bool(wallet.btc_wallet),
            bool(wallet.eth_wallet),
            bool(wallet.usdt_erc20_wallet),
            bool(wallet.usdt_trc20_wallet),
        ])

    context = {
        "wallet_exists": wallet_exists,
        "wallet_has_address": wallet_has_address,
        "profile": profile,
    }

    return render(
        request,
        "users/deposit.html",
        context
    )

@login_required
def make_deposit(request):
    profile = UserProfile.objects.get(user=request.user)

    # ---------------------------------------------------------
    # 1. USER MUST HAVE A USERWALLET RECORD
    # ---------------------------------------------------------
    wallet = UserWallet.objects.filter(
        user=request.user
    ).first()

    if not wallet:
        messages.error(
            request,
            "⚠️ Please update your wallet before making deposits."
        )
        return redirect("users:deposit")

    # ---------------------------------------------------------
    # 2. USER MUST BE VERIFIED
    # ---------------------------------------------------------
    verification = UserVerification.objects.filter(
        user=request.user
    ).first()

    if not verification or not verification.is_verified:
        messages.error(
            request,
            "⚠️ Your account is not verified. Please wait for admin approval before making deposits."
        )
        return redirect("users:deposit")

    # ---------------------------------------------------------
    # 3. ONLY ACCEPT POST REQUESTS
    # ---------------------------------------------------------
    if request.method != "POST":
        return redirect("users:deposit")

    # ---------------------------------------------------------
    # 4. GET AMOUNT
    # ---------------------------------------------------------
    try:
        amount = Decimal(request.POST.get("amount"))
    except (TypeError, ValueError, ArithmeticError):
        messages.error(
            request,
            "Invalid amount entered."
        )
        return redirect("users:deposit")

    # ---------------------------------------------------------
    # 5. GET CURRENCY
    # ---------------------------------------------------------
    currency = request.POST.get("currency")

    # ---------------------------------------------------------
    # 6. CHECK THE WALLET FOR THE SPECIFIC CURRENCY
    # ---------------------------------------------------------
    wallet_address = None

    if currency == "BTC":
        wallet_address = wallet.btc_wallet

    elif currency == "ETH":
        wallet_address = wallet.eth_wallet

    elif currency == "USDT ERC20":
        wallet_address = wallet.usdt_erc20_wallet

    elif currency == "USDT TRC20":
        wallet_address = wallet.usdt_trc20_wallet

    elif currency == "USD":
        # USD is a balance in your system, not one of
        # the four wallet addresses.
        wallet_address = "USD"

    else:
        messages.error(
            request,
            "Invalid currency selected."
        )
        return redirect("users:deposit")

    # ---------------------------------------------------------
    # 7. REQUIRE THE CORRECT WALLET ADDRESS
    # ---------------------------------------------------------
    if currency != "USD" and not wallet_address:
        messages.error(
            request,
            f"⚠️ Please update your {currency} wallet before making this deposit."
        )
        return redirect("users:deposit")

    # ---------------------------------------------------------
    # 8. MINIMUM DEPOSIT
    # ---------------------------------------------------------
    if amount < MIN_DEPOSIT:
        messages.error(
            request,
            f"The minimum deposit is ${MIN_DEPOSIT}. Please enter a valid amount."
        )
        return redirect("users:deposit")

    # ---------------------------------------------------------
    # 9. KEEP YOUR EXISTING BALANCE BEHAVIOR
    # ---------------------------------------------------------
    if currency == "BTC":
        profile.btc_balance += amount

    elif currency == "ETH":
        profile.eth_balance += amount

    elif currency == "USDT ERC20":
        profile.usdt_erc20_balance += amount

    elif currency == "USDT TRC20":
        profile.usdt_trc20_balance += amount

    elif currency == "USD":
        profile.usd_balance += amount

    profile.save()

    # ---------------------------------------------------------
    # 10. CREATE THE PENDING DEPOSIT
    # ---------------------------------------------------------
    deposit = Deposit.objects.create(
        user=request.user,
        amount=amount,
        currency=currency,
        status="pending"
    )

    # ---------------------------------------------------------
    # 11. SEND USER TO THE INVOICE
    # ---------------------------------------------------------
    return redirect(
        "users:deposit_invoice",
        deposit_id=deposit.id,
        currency=currency
    )

@login_required
def deposit_invoice(request, deposit_id, currency):
    deposit = get_object_or_404(Deposit, id=deposit_id, user=request.user)
    company_wallet = CompanyWallet.objects.first()

    if not company_wallet:
        messages.error(request, "No company wallet configured. Please contact support.")
        return redirect("users:deposit")

    wallet_address, qr_code_url = None, None

    if currency == "BTC":
        wallet_address = company_wallet.btc_wallet
        qr_code_url = company_wallet.btc_qr.url if company_wallet.btc_qr else None
    elif currency == "ETH":
        wallet_address = company_wallet.eth_wallet
        qr_code_url = company_wallet.eth_qr.url if company_wallet.eth_qr else None
    elif currency == "USDT ERC20":
        wallet_address = company_wallet.usdt_erc20_wallet
        qr_code_url = company_wallet.usdt_erc20_qr.url if company_wallet.usdt_erc20_qr else None
    elif currency == "USDT TRC20":
        wallet_address = company_wallet.usdt_trc20_wallet
        qr_code_url = company_wallet.usdt_trc20_qr.url if company_wallet.usdt_trc20_qr else None

    if not wallet_address:
        messages.error(request, f"No {currency} wallet configured. Please contact support.")
        return redirect("users:deposit")

    return render(request, "users/deposit_invoice.html", {
        "deposit": deposit,
        "currency": currency,
        "wallet_address": wallet_address,
        "qr_code_url": qr_code_url,
    })


@login_required
def request_withdrawal(request):
    profile = UserProfile.objects.get(user=request.user)
    userwallet = UserWallet.objects.filter(user=request.user).first()

    # =========================================================
    # POST REQUEST
    # =========================================================
    if request.method == "POST":

        # =====================================================
        # VERIFICATION CHECK
        # Only verified users can make withdrawals.
        # IMPORTANT:
        # This check is ONLY performed when the user submits
        # a withdrawal. It must not redirect GET requests
        # back to this same URL.
        # =====================================================
        if profile.verification_status != "verified":
            messages.error(
                request,
                "⚠️ Your account is not verified. "
                "Please wait for admin approval before making withdrawals."
            )
            return redirect("users:request_withdrawal")

        form = WithdrawalForm(request.POST)

        if not form.is_valid():
            messages.error(
                request,
                "Please complete all required fields."
            )
            return redirect("users:request_withdrawal")

        # =====================================================
        # AMOUNT
        # =====================================================
        try:
            amount = Decimal(str(form.cleaned_data["amount"]))
        except InvalidOperation:
            messages.error(
                request,
                "Invalid withdrawal amount."
            )
            return redirect("users:request_withdrawal")

        currency = form.cleaned_data["currency"]
        wallet_address = form.cleaned_data["wallet_address"].strip()

        # =====================================================
        # BASIC VALIDATION
        # =====================================================
        if amount <= 0:
            messages.error(
                request,
                "Withdrawal amount must be greater than zero."
            )
            return redirect("users:request_withdrawal")

        if not wallet_address:
            messages.error(
                request,
                "Destination wallet address is required."
            )
            return redirect("users:request_withdrawal")

        # =====================================================
        # BALANCE CHECKS
        # =====================================================
        if currency == "BTC":

            if profile.btc_balance < amount:
                messages.error(
                    request,
                    "Insufficient BTC balance."
                )
                return redirect("users:request_withdrawal")

            profile.btc_balance -= amount

        elif currency == "ETH":

            if profile.eth_balance < amount:
                messages.error(
                    request,
                    "Insufficient ETH balance."
                )
                return redirect("users:request_withdrawal")

            profile.eth_balance -= amount

        elif currency == "USDT_ERC20":

            if profile.usdt_erc20_balance < amount:
                messages.error(
                    request,
                    "Insufficient USDT ERC20 balance."
                )
                return redirect("users:request_withdrawal")

            profile.usdt_erc20_balance -= amount

        elif currency == "USDT_TRC20":

            if profile.usdt_trc20_balance < amount:
                messages.error(
                    request,
                    "Insufficient USDT TRC20 balance."
                )
                return redirect("users:request_withdrawal")

            profile.usdt_trc20_balance -= amount

        else:

            messages.error(
                request,
                "Invalid currency selected."
            )
            return redirect("users:request_withdrawal")

        # =====================================================
        # SAVE ATOMICALLY
        # =====================================================
        with transaction.atomic():

            profile.save()

            withdrawal = Withdrawal.objects.create(
                user=request.user,
                currency=currency,
                amount=amount,
                wallet_address=wallet_address,
                status="pending",
            )

            Transaction.objects.create(
                user=request.user,
                asset=currency,
                amount=amount,
                type="withdrawal",
                status="pending",
            )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================
        messages.success(
            request,
            f"Your withdrawal request of {amount} {currency} "
            f"has been submitted successfully."
        )

        # =====================================================
        # WITHDRAWAL INVOICE
        # =====================================================
        return redirect(
            "users:withdraw_invoice",
            withdrawal_id=withdrawal.id,
            currency=currency
        )

    # =========================================================
    # GET REQUEST
    # =========================================================
    form = WithdrawalForm()

    # ---------------------------------------------------------
    # Prevent errors if the user has not created a wallet yet.
    # ---------------------------------------------------------
    wallets = {
        "BTC": userwallet.btc_wallet if userwallet else "",

        "ETH": userwallet.eth_wallet if userwallet else "",

        "USDT_ERC20": (
            userwallet.usdt_erc20_wallet
            if userwallet else ""
        ),

        "USDT_TRC20": (
            userwallet.usdt_trc20_wallet
            if userwallet else ""
        ),
    }

    return render(
        request,
        "users/withdraw.html",
        {
            "form": form,
            "profile": profile,
            "userwallet": userwallet,
            "wallets": wallets,
        }
    )


@login_required
def withdraw_invoice(request, withdrawal_id, currency):
    withdrawal = get_object_or_404(Withdrawal, id=withdrawal_id, user=request.user)
    userwallet = UserWallet.objects.get(user=request.user)
    wallets = {
        "BTC": userwallet.btc_wallet,
        "ETH": userwallet.eth_wallet,
        "USDT_ERC20": userwallet.usdt_erc20_wallet,
        "USDT_TRC20": userwallet.usdt_trc20_wallet,
    }
    wallet_address = wallets.get(currency, "N/A")

    return render(request, "users/withdraw_invoice.html", {
        "withdrawal": withdrawal,
        "currency": currency,
        "wallet_address": wallet_address,
        "userwallet": userwallet, 
    })




@login_required
def manage_wallets(request):
    wallet, created = UserWallet.objects.get_or_create(user=request.user)

    if request.method == "POST":
        form = UserWalletForm(request.POST, instance=wallet)
        if form.is_valid():
            form.save()
            return redirect("users:make_withdrawal")
    else:
        form = UserWalletForm(instance=wallet)

    return render(request, "users/manage_wallets.html", {
        "form": form
    })


@login_required
def start_investment(request):
    plans = InvestmentPlan.objects.all()

    profile = get_object_or_404(
        UserProfile,
        user=request.user
    )

    if request.method == "POST":
        plan = get_object_or_404(
            InvestmentPlan,
            id=request.POST.get("plan_id")
        )

        try:
            amount = Decimal(request.POST.get("amount"))
        except Exception:
            messages.error(request, "Please enter a valid investment amount.")
            return redirect("users:investment_plans")

        # Check investment limits
        if amount < plan.min_amount or amount > plan.max_amount:
            messages.error(
                request,
                f"Amount must be between ${plan.min_amount} and ${plan.max_amount}."
            )
            return redirect("users:investment_plans")

        # Check wallet balance
        if profile.usd_balance < amount:
            messages.error(request, "Insufficient wallet balance.")
            return redirect("users:investment_plans")

        # Deduct funds from wallet
        profile.usd_balance -= amount

        # Move funds into investment balance
        profile.investment_balance += amount

        profile.save()

        # Investment dates
        start_date = timezone.now()
        end_date = start_date + timedelta(hours=plan.duration_hours)

        # Create investment
        ActiveInvestment.objects.create(
            user=request.user,
            plan_name=plan.name,
            roi_percent=plan.roi_percent,
            amount=amount,
            start_date=start_date,
            end_date=end_date,
            status="active"
        )

        # Friendly duration
        if plan.duration_hours % 24 == 0:
            duration = f"{plan.duration_hours // 24} day(s)"
        else:
            duration = f"{plan.duration_hours} hour(s)"

        messages.success(
            request,
            f"You successfully invested ${amount} in {plan.name}. "
            f"Duration: {duration}. ROI: {plan.roi_percent}%."
        )

        return redirect("users:investments")

    return render(
        request,
        "users/start_investment.html",
        {
            "plans": plans,
            "profile": profile,
        },
    )


@login_required
def manage_2fa(request):
    if not request.session.get("twofa_secret"):
        request.session["twofa_secret"] = pyotp.random_base32()

    secret = request.session["twofa_secret"]

    otp_uri = pyotp.TOTP(secret).provisioning_uri(
        name=f"{request.user.username} - BitnetFX",  
        issuer_name="BitnetFX"
    )

    qr = qrcode.make(otp_uri)
    buf = io.BytesIO()
    qr.save(buf, format="PNG")
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    if request.method == "POST":
        code = request.POST.get("otp")
        totp = pyotp.TOTP(secret)

        # Commented out the enforcement of 2FA verification
        # if totp.verify(code):
        #     verification, _ = UserVerification.objects.get_or_create(
        #         user=request.user
        #     )
        #     verification.secret = secret   # 👈 save the secret permanently
        #     verification.is_verified = True
        #     verification.save()
        #
        #     # Clear session secret after successful setup
        #     del request.session["twofa_secret"]
        #
        #     messages.success(
        #         request,
        #         "Two-Factor Authentication enabled successfully."
        #     )
        #     return redirect("users:security")
        #
        # messages.error(
        #     request,
        #     "Invalid code, please try again."
        # )

      
        verification, _ = UserVerification.objects.get_or_create(user=request.user)
        verification.secret = secret
        verification.is_verified = True
        verification.save()

       
        del request.session["twofa_secret"]

        messages.success(request, "2FA secret saved with username label.")
        return redirect("users:security")

    return render(request, "users/manage_2fa.html", {
        "qr_b64": qr_b64
    })


@login_required
def verify_identity(request):
    if request.method == "POST":
        doc = request.FILES.get("document")

        kyc, _ = UserKYC.objects.get_or_create(
            user=request.user
        )

        kyc.document = doc
        kyc.status = "pending"
        kyc.save()

        return redirect("users:security_center")

    return render(
        request,
        "users/verify_identity.html",
        {
            "kyc": UserKYC.objects.filter(user=request.user).first()
        }
    )


@login_required
def security_center(request):
    user = request.user

    verification, _ = UserVerification.objects.get_or_create(user=user)
    recovery_codes = RecoveryCode.objects.filter(user=user, used=False).values_list("code", flat=True)
    profile = UserProfile.objects.get(user=user)

    # Normalize verification status
    kyc_status = profile.verification_status.lower() if profile.verification_status else "pending"

    recent_logins = []

    return render(
        request,
        "users/security.html",
        {
            "verification": verification,
            "qr_code": None,
            "recovery_codes": recovery_codes,
            "kyc_status": kyc_status,
            "recent_logins": recent_logins,
        }
    )

@login_required
def manage_api_keys(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "generate":
            APIKey.objects.create(user=request.user)
            messages.success(request, "New API key generated.")
        elif action == "revoke":
            key_id = request.POST.get("key_id")
            key = APIKey.objects.filter(id=key_id, user=request.user).first()
            if key:
                key.status = "Revoked"
                key.save()
                messages.success(request, "API key revoked.")
        return redirect("users:manage_api_keys")

    api_keys = APIKey.objects.filter(user=request.user)
    return render(request, "users/manage_api_keys.html", {"api_keys": api_keys})


    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="portfolio.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Type', 'Asset', 'Amount'])

    transactions = Transaction.objects.filter(user=request.user).order_by('-date')
    for tx in transactions:
        writer.writerow([tx.date.strftime("%Y-%m-%d %H:%M"), tx.type, tx.asset, tx.amount])

    return response


@login_required
def export_portfolio_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="portfolio.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 800, "Portfolio Transactions")

    y = 760
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')
    for tx in transactions:
        line = f"{tx.date.strftime('%Y-%m-%d %H:%M')} - {tx.type} - {tx.asset} - ${tx.amount}"
        p.drawString(100, y, line)
        y -= 20

    p.showPage()
    p.save()
    return response


@login_required
def export_portfolio_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="portfolio.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Type', 'Asset', 'Amount'])

    transactions = Transaction.objects.filter(user=request.user).order_by('-date')
    for tx in transactions:
        writer.writerow([tx.date.strftime("%Y-%m-%d %H:%M"), tx.type, tx.asset, tx.amount])

    return response


login_required
def export_portfolio_excel(request):
  
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio Report"

   
    investments = ActiveInvestment.objects.filter(user=request.user)
    total_invested = investments.aggregate(Sum('amount'))['amount__sum'] or 0
    current_value = sum(inv.amount for inv in investments if inv.status == "active")
    roi = ((current_value - total_invested) / total_invested * 100) if total_invested else 0
    best_asset = investments.first().plan_name if investments.exists() else "N/A"

    ws.append([f"{request.user.username}'s Portfolio Report"])
    ws.append([])
    ws.append(["Total Invested", total_invested])
    ws.append(["Current Value", round(current_value, 2)])
    ws.append(["ROI (%)", round(roi, 2)])
    ws.append(["Best Asset", best_asset])
    ws.append([])

  
    ws.append(["Date", "Type", "Asset", "Amount"])
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')
    for tx in transactions:
        ws.append([tx.date.strftime("%Y-%m-%d %H:%M"), tx.type, tx.asset, tx.amount])

  
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="portfolio.xlsx"'
    wb.save(response)
    return response


@login_required
def profit_history(request):
    user = request.user
    profits = ProfitRecord.objects.filter(user=user).order_by("-date")

    total_profit = profits.aggregate(total=Sum("amount"))["total"] or 0
    monthly_profit = profits.filter(date__month=timezone.now().month).aggregate(total=Sum("amount"))["total"] or 0

    chart_labels = [p.date.strftime("%b %d") for p in profits]
    chart_data = [float(p.amount) for p in profits]

    context = {
        "profits": profits,
        "total_profit": total_profit,
        "monthly_profit": monthly_profit,
        "chart_labels": json.dumps(chart_labels),
        "chart_data": json.dumps(chart_data),
    }
    return render(request, "users/profit_history.html", context)

@login_required
def complete_investment(request, investment_id):
    investment = get_object_or_404(ActiveInvestment, id=investment_id, user=request.user)

    if investment.status == "active":
        credit_profit(request.user, investment)   
        investment.status = "completed"
        investment.save()

    return redirect("users:profit_history")

def award_referral_commission(deposit):
    referral_user = deposit.user
    try:
        referral_record = Referral.objects.get(user=referral_user)
        referrer = referral_record.referrer

        if referrer:
            deposit_amount = deposit.amount
            commission_amount = deposit_amount * Decimal("0.07")  # 7%

            ReferralCommission.objects.create(
                referrer=referrer,
                referral=referral_user,
                deposit_amount=deposit_amount,
                commission_amount=commission_amount
            )

           
            referrer.wallet.balance += commission_amount
            referrer.wallet.save()

    except Referral.DoesNotExist:
       
        pass


@login_required
def ai_chat(request):
    if request.method == "POST":

        # User wallet
        wallet = Wallet.objects.filter(user=request.user).first()

        # Saved wallet addresses
        user_wallet = UserWallet.objects.filter(user=request.user).first()

        # Company fallback wallets
        company_wallet = CompanyWallet.objects.first()

        ai_response = "🤖 I'm your AI Trading Assistant."

        # ===============================
        # AI Investment (only if AI invests)
        # ===============================
        if "plan_id" in request.POST:

            plan_id = request.POST.get("plan_id")
            amount = Decimal(request.POST.get("amount", "0"))

            plan = InvestmentPlan.objects.filter(id=plan_id).first()

            if plan and wallet and plan.min_amount <= amount <= plan.max_amount:

                if wallet.balance >= amount:

                    wallet.balance -= amount
                    wallet.save()

                    Transaction.objects.create(
                        user=request.user,
                        asset="USD",
                        amount=amount,
                        type="investment",
                        status="pending",
                    )

                    ai_response = (
                        f"✅ You invested ${amount} in "
                        f"{plan.name}. ROI: {plan.roi_percent}% "
                        f"{plan.duration_text}."
                    )

                else:
                    ai_response = "❌ Insufficient balance to invest."

            else:
                ai_response = "❌ Invalid investment plan or amount."

            return JsonResponse({
                "response": ai_response
            })

        # ===============================
        # Chat Messages
        # ===============================

        user_message = request.POST.get("message", "").lower()

        # -------------------------------
        # Balance
        # -------------------------------
        if "balance" in user_message:

            if wallet:
                ai_response = (
                    f"Your wallet balance is "
                    f"${wallet.balance}. "
                    f"Gas fees paid: ${wallet.gas_fee_paid}."
                )
            else:
                ai_response = (
                    "You don't have a wallet yet."
                )

        # -------------------------------
        # Deposit Instructions
        # -------------------------------
        elif "deposit" in user_message:

            if "btc" in user_message:

                address = ""

                if user_wallet and user_wallet.btc_wallet:
                    address = user_wallet.btc_wallet
                elif company_wallet:
                    address = company_wallet.btc_wallet

                ai_response = (
                    f"Bitcoin (BTC)\n"
                    f"Wallet Address: {address}\n"
                    f"Minimum Deposit: 0.0005 BTC\n"
                    f"Confirmations: 2\n"
                    f"⚠ Send only BTC to this address."
                )

            elif "eth" in user_message:

                address = ""

                if user_wallet and user_wallet.eth_wallet:
                    address = user_wallet.eth_wallet
                elif company_wallet:
                    address = company_wallet.eth_wallet

                ai_response = (
                    f"Ethereum (ETH)\n"
                    f"Wallet Address: {address}\n"
                    f"Minimum Deposit: 0.01 ETH\n"
                    f"Confirmations: 12\n"
                    f"⚠ Send only ETH to this address."
                )

            elif "erc20" in user_message:

                address = ""

                if user_wallet and user_wallet.usdt_erc20_wallet:
                    address = user_wallet.usdt_erc20_wallet
                elif company_wallet:
                    address = company_wallet.usdt_erc20_wallet

                ai_response = (
                    f"USDT ERC20\n"
                    f"Wallet Address: {address}\n"
                    f"Minimum Deposit: 50 USDT\n"
                    f"Confirmations: 20\n"
                    f"⚠ Send only USDT ERC20 to this address."
                )

            elif "trc20" in user_message:

                address = ""

                if user_wallet and user_wallet.usdt_trc20_wallet:
                    address = user_wallet.usdt_trc20_wallet
                elif company_wallet:
                    address = company_wallet.usdt_trc20_wallet

                ai_response = (
                    f"USDT TRC20\n"
                    f"Wallet Address: {address}\n"
                    f"Minimum Deposit: 50 USDT\n"
                    f"Confirmations: 20\n"
                    f"⚠ Send only USDT TRC20 to this address."
                )

            else:

                ai_response = (
                    "Please specify the asset you want to deposit "
                    "(BTC, ETH, ERC20 or TRC20)."
                )

        # -------------------------------
        # Transaction History
        # -------------------------------
        elif "history" in user_message or "transactions" in user_message:

            transactions = Transaction.objects.filter(
                user=request.user
            ).order_by("-date")[:5]

            if transactions:

                history = "\n".join([
                    f"{t.type.capitalize()} "
                    f"{t.amount} "
                    f"{t.asset} "
                    f"- {t.status}"
                    for t in transactions
                ])

                ai_response = (
                    f"Recent transactions:\n{history}"
                )

            else:

                ai_response = "No transactions found."

        # -------------------------------
        # Plan Recommendation
        # -------------------------------
        elif "plan" in user_message or "recommend" in user_message:

            if wallet:

                balance = wallet.balance

                plan = InvestmentPlan.objects.filter(
                    min_amount__lte=balance,
                    max_amount__gte=balance,
                ).first()

                if plan:

                    ai_response = (
                        f"Based on your balance "
                        f"(${balance}), I recommend "
                        f"{plan.name} "
                        f"({plan.roi_percent}% ROI "
                        f"after {plan.duration_text})."
                    )

                else:

                    ai_response = (
                        "No suitable investment plan "
                        "matches your balance."
                    )

            else:

                ai_response = (
                    "You need a wallet before "
                    "I can recommend a plan."
                )

        # -------------------------------
        # Withdraw
        # -------------------------------
        elif "withdraw" in user_message:

            ai_response = (
                "Use the withdrawal form below. "
                "When you select an asset, "
                "your saved wallet address "
                "will automatically appear."
            )

        # -------------------------------
        # Future Trading
        # -------------------------------
        else:

            trade_match = re.match(
                r"(buy|sell)\s+([\d\.]+)\s*(btc|eth|usdt)?",
                user_message
            )

            if trade_match:

                action, amount, asset = trade_match.groups()

                amount = Decimal(amount)

                asset = asset.upper() if asset else "BTC"

                ai_response = (
                    f"⚠ Trade feature coming soon:\n"
                    f"{action.capitalize()} "
                    f"{amount} {asset}"
                )

        return JsonResponse({
            "response": ai_response
        })

    # ==================================
    # GET Request
    # ==================================

    wallet = Wallet.objects.filter(user=request.user).first()

    user_wallet = UserWallet.objects.filter(
        user=request.user
    ).first()

    transactions = Transaction.objects.filter(
        user=request.user
    ).order_by("-date")[:5]

    plans = InvestmentPlan.objects.all()

    recommended_plan = None

    if wallet:

        recommended_plan = InvestmentPlan.objects.filter(
            min_amount__lte=wallet.balance,
            max_amount__gte=wallet.balance,
        ).first()

    return render(
        request,
        "users/ai_chat.html",
        {
            "wallet": wallet,
            "user_wallet": user_wallet,
            "transactions": transactions,
            "plans": plans,
            "recommended_plan": recommended_plan,
        },
    )


@login_required
def wallet_data(request):
    wallet = Wallet.objects.filter(user=request.user).first()
    user_wallet = UserWallet.objects.filter(user=request.user).first()

    if not wallet or not user_wallet:
        return JsonResponse({"error": "Wallet not found"}, status=404)

    return JsonResponse({
        "balance": str(wallet.balance),
        "gas_fee_paid": str(wallet.gas_fee_paid),

        "btc_balance": str(user_wallet.btc_balance),
        "eth_balance": str(user_wallet.eth_balance),

        "btc_wallet": user_wallet.btc_wallet,
        "eth_wallet": user_wallet.eth_wallet,
        "usdt_erc20_wallet": user_wallet.usdt_erc20_wallet,
        "usdt_trc20_wallet": user_wallet.usdt_trc20_wallet,
    })


@login_required
def recent_transactions_api(request):
    transactions = Transaction.objects.filter(user=request.user).order_by("-date")[:10]
    data = [
        {
            "type": tx.type,
            "amount": float(tx.amount),
            "asset": tx.asset,
            "status": tx.status,
            "date": tx.date.strftime("%Y-%m-%d %H:%M:%S"),
        }
        for tx in transactions
    ]
    return JsonResponse(data, safe=False)


@login_required
def wallet_api(request):
    wallet = Wallet.objects.filter(user=request.user).first()
    if not wallet:
        return JsonResponse({"error": "No wallet found"}, status=404)

    data = {
        "balance": float(wallet.balance),
        "gas_fee_paid": float(wallet.gas_fee_paid),
        "btc_balance": float(wallet.userwallet.btc_balance),
        "eth_balance": float(wallet.userwallet.eth_balance),
        "usdt_erc20_wallet": float(wallet.userwallet.usdt_erc20_wallet),
        "usdt_trc20_wallet": float(wallet.userwallet.usdt_trc20_wallet),
    }
    return JsonResponse(data)

@login_required
def company_wallet(request, asset):
    wallet = CompanyWallet.objects.first()

    if not wallet:
        return JsonResponse({"error": "Company wallet not configured."}, status=404)

    asset = asset.lower()

    if asset == "btc":
        return JsonResponse({
            "asset": "Bitcoin (BTC)",
            "address": wallet.btc_wallet,
            "qr": wallet.btc_qr.url if wallet.btc_qr else "",
            "network": "Bitcoin",
            "minimum": "0.0005 BTC",
            "confirmations": "2",
        })

    elif asset == "eth":
        return JsonResponse({
            "asset": "Ethereum (ETH)",
            "address": wallet.eth_wallet,
            "qr": wallet.eth_qr.url if wallet.eth_qr else "",
            "network": "ERC20",
            "minimum": "0.01 ETH",
            "confirmations": "12",
        })

    elif asset == "erc20":
        return JsonResponse({
            "asset": "USDT ERC20",
            "address": wallet.usdt_erc20_wallet,
            "qr": wallet.usdt_erc20_qr.url if wallet.usdt_erc20_qr else "",
            "network": "ERC20",
            "minimum": "10 USDT",
            "confirmations": "12",
        })

    elif asset == "trc20":
        return JsonResponse({
            "asset": "USDT TRC20",
            "address": wallet.usdt_trc20_wallet,
            "qr": wallet.usdt_trc20_qr.url if wallet.usdt_trc20_qr else "",
            "network": "TRC20",
            "minimum": "10 USDT",
            "confirmations": "20",
        })

    return JsonResponse({"error": "Invalid asset."}, status=400)

